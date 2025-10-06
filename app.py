#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cybear404 AutoCut Optimizer
---------------------------------
A free proof-of-concept tool for optimizing material cuts with minimal waste.

Features:
- Modern GUI (PySide6 / Qt) with Cybear404 branding
  (light/dark theme toggle is experimental and not fully working yet)
- Reads/writes Excel (.xlsx) with pandas + openpyxl
- Bin-packing–style grouping by stock length and material
- Configurable saw kerf (fractional or decimal input)
- Options to overwrite or copy workbook
- Optional reports: Summary, Validation, Waste Report, Procurement, Issues
- Sample data/template generator for quick testing
- Settings persistence and "Open Output File" button

Intended use:
- Demonstration / proof-of-concept for construction, fabrication, and shop workflows
- Free to download and use for testing/learning
- Not certified for production-critical environments (use at your own risk)

Author: Cybear404, LLC
Contributors: Developed with assistance from OpenAI's ChatGPT
URL: https://cybear404.com

License:
- Noncommercial use only under PolyForm Noncommercial 1.0.0 (see LICENSE)
- Commercial use requires a separate license (see COMMERCIAL-LICENSE.md)
Branding:
- Cybear404 name/logo are proprietary (see BRANDING.md)
"""

from __future__ import annotations

import sys
import os
import json
import platform
import random
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QAction, QIcon, QPalette, QColor, QFont, QPixmap
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox, QLabel,
    QVBoxLayout, QHBoxLayout, QFormLayout, QLineEdit, QPushButton, QCheckBox,
    QTextEdit, QSpacerItem, QSizePolicy, QFrame
)

# ----------------------------- Version & paths ------------------------------

def resource_path(relative_path: str) -> str:
    """Return absolute path to a bundled resource.

    Works in dev (uses file dir) and in PyInstaller onefile/onedir builds
    (uses the temporary _MEIPASS folder that PyInstaller extracts to).
    """
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    return str(base / relative_path)


def _read_version(default: str = "0.1.0") -> str:
    try:
        vpath = Path(resource_path("VERSION"))
        if vpath.exists():
            v = vpath.read_text(encoding="utf-8").strip()
            return v or default
    except Exception:
        pass
    return default

__version__ = _read_version("0.1.0")

APP_TITLE = "Cybear404 AutoCut Optimizer"
DEFAULT_SHEET_NAME = "Grouped Cuts"
TEMPLATE_BASENAME = "AUTOCUT_GROUP_GENERATOR_Template"
VENDOR_DIR = os.path.join(os.path.expanduser("~"), ".cybear404_autocut_optimizer")
SETTINGS_PATH = os.path.join(VENDOR_DIR, "settings.json")

# ----------------------------- Utility -------------------------------------

def parse_fraction_or_decimal(text: str) -> float:
    t = (text or "").strip()
    if not t:
        raise ValueError("Kerf width is required.")
    try:
        if "/" in t:
            num, den = t.split("/", 1)
            return float(num) / float(den)
        return float(t)
    except Exception:
        raise ValueError("Invalid kerf format. Use decimal like 0.125 or fraction like 1/8.")


def autosize_openpyxl_columns(ws) -> None:
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is not None:
                max_length = max(max_length, len(str(v)))
        ws.column_dimensions[col_letter].width = max_length + 2

# ---------------------------- Core Logic ------------------------------------

Bins = List[List[float]]
GroupedKey = Tuple[str, str, float]
GroupedBins = Dict[GroupedKey, Bins]


def bin_packing_grouped(df: pd.DataFrame, saw_width: float, *, allow_split_oversize: bool) -> Tuple[GroupedBins, List[dict]]:
    grouped_bins: GroupedBins = {}
    issues: List[dict] = []

    for (max_length, material_type), group in df.groupby(["Max Length", "Material Type"], dropna=False):
        items: List[float] = []
        for x in group["Cut Length"].tolist():
            try:
                items.append(float(x))
            except Exception:
                pass

        items.sort(reverse=True)
        bins: Bins = []
        try:
            max_len_val = float(max_length)
        except Exception:
            continue

        for item in items:
            remaining_item = item
            if remaining_item > max_len_val and not allow_split_oversize:
                issues.append({
                    "Material": str(material_type),
                    "Stock Length": float(max_len_val),
                    "Cut Length": float(remaining_item),
                    "Problem": "Cut length exceeds stock; cannot be produced as a single piece."
                })
                continue
            while remaining_item > max_len_val and allow_split_oversize:
                bins.append([max_len_val])
                remaining_item -= max_len_val

            placed = False
            best_idx = -1
            least_waste = float('inf')

            for idx, b in enumerate(bins):
                used = sum(b) + saw_width * len(b)
                remaining_len = max_len_val - used
                needed = remaining_item + (saw_width if b else 0.0)
                if remaining_len >= needed:
                    waste_after = remaining_len - needed
                    if waste_after < least_waste:
                        best_idx = idx
                        least_waste = waste_after

            if best_idx != -1:
                bins[best_idx].append(remaining_item)
                placed = True

            if not placed:
                bins.append([remaining_item])

        label = f"{int(max_len_val) if float(max_len_val).is_integer() else max_len_val} {material_type}"
        grouped_bins[(label, str(material_type), float(max_len_val))] = bins

    return grouped_bins, issues


def write_grouped_bins(file_path: str, grouped_bins: GroupedBins, *,
                       base_sheet_name: str = DEFAULT_SHEET_NAME,
                       overwrite_sheet: bool = False,
                       write_to_copy: bool = False) -> str:
    out_path = file_path
    if write_to_copy:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        root, ext = os.path.splitext(file_path)
        out_path = f"{root}_Cybear404_{ts}{ext}"
        wb_src = load_workbook(file_path)
        wb_src.save(out_path)

    wb = load_workbook(out_path)

    sheet_name = base_sheet_name
    if sheet_name in wb.sheetnames:
        if overwrite_sheet:
            del wb[sheet_name]
        else:
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_sheet_name} {counter}"
                counter += 1

    ws = wb.create_sheet(sheet_name)

    output_data: Dict[str, List[str]] = {}
    for (group_label, material_type, L), bins in grouped_bins.items():
        lines: List[str] = []
        for i, b in enumerate(sorted(bins, key=lambda x: -sum(x))):
            pretty = [int(x) if float(x).is_integer() else round(float(x), 3) for x in b]
            lines.append(f"Group {i+1}: {pretty}")
        output_data[group_label] = lines

    col_idx = 1
    for group_label, lines in output_data.items():
        ws.cell(row=1, column=col_idx, value=group_label)
        for r, line in enumerate(lines, start=2):
            ws.cell(row=r, column=col_idx, value=line)
        col_idx += 1

    autosize_openpyxl_columns(ws)
    wb.save(out_path)
    wb.close()
    return out_path

# ----------- Optional Sheets: Summary, Validation, Waste Report, Procurement, Issues -------------

def build_summary(grouped_bins: GroupedBins, kerf: float) -> pd.DataFrame:
    rows = []
    for (label, mat, L), bins in grouped_bins.items():
        cuts_total = sum(sum(b) for b in bins)
        kerf_total = sum(kerf * len(b) for b in bins if b)
        waste_total = sum(max(0.0, L - (sum(b) + kerf * len(b))) for b in bins if b)
        bars = len(bins)
        stock_total = L * bars
        util = (stock_total - waste_total) / stock_total if stock_total else 0
        rows.append({
            "Group": label,
            "Material": mat,
            "Stock Length": L,
            "Bars Used": bars,
            "Total Cut Length": round(cuts_total, 3),
            "Approx Kerf Used": round(kerf_total, 3),
            "Estimated Waste": round(waste_total, 3),
            "Utilization %": round(util * 100, 2),
        })
    return pd.DataFrame(rows).sort_values(["Material", "Stock Length"]) if rows else pd.DataFrame()


def write_summary_sheet(wb_path: str, df_summary: pd.DataFrame) -> None:
    wb = load_workbook(wb_path)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")
    if not df_summary.empty:
        ws.append(list(df_summary.columns))
        for _, row in df_summary.iterrows():
            ws.append([row[c] for c in df_summary.columns])
    autosize_openpyxl_columns(ws)
    wb.save(wb_path)


def write_validation_sheet(wb_path: str, grouped_bins: GroupedBins, kerf: float, df_summary: pd.DataFrame) -> None:
    feasible_violations = 0
    for (label, mat, L), bins in grouped_bins.items():
        for b in bins:
            used = sum(b) + kerf * len(b)
            if used - L > 1e-6:
                feasible_violations += 1

    problems = []
    if feasible_violations:
        problems.append(f"{feasible_violations} bar(s) exceed stock length feasibility.")

    if "Estimated Waste" in df_summary.columns and (df_summary["Estimated Waste"] < -1e-6).any():
        problems.append("Negative waste detected (should not happen).")

    wb = load_workbook(wb_path)
    if "Validation" in wb.sheetnames:
        del wb["Validation"]
    ws = wb.create_sheet("Validation")
    ws.append(["Check", "Result"])
    ws.append(["Per-bin feasibility (<= stock length)", "PASS" if feasible_violations == 0 else "FAIL"])
    ws.append(["Non-negative waste", "PASS" if not problems else ("FAIL" if any("Negative waste" in p for p in problems) else "PASS")])
    ws.append(["Notes", "; ".join(problems) if problems else "All checks passed."])
    autosize_openpyxl_columns(ws)
    wb.save(wb_path)


def write_waste_report_sheet(wb_path: str, grouped_bins: GroupedBins, kerf: float) -> None:
    wb = load_workbook(wb_path)
    if "Waste Report" in wb.sheetnames:
        del wb["Waste Report"]
    ws = wb.create_sheet("Waste Report")
    ws.append(["Group", "Material", "Stock Length", "Bar #", "Pieces", "Sum Cuts", "Kerf Used", "Used Total", "Leftover", "Util % (bar)"])
    for (label, mat, L), bins in grouped_bins.items():
        for i, b in enumerate(sorted(bins, key=lambda x: -sum(x)), start=1):
            sum_cuts = sum(b)
            kerf_used = kerf * len(b)
            used_total = sum_cuts + kerf_used
            leftover = max(0.0, L - used_total)
            util = used_total / L * 100 if L else 0
            pretty = [int(x) if float(x).is_integer() else round(float(x), 3) for x in b]
            ws.append([label, mat, L, i, str(pretty), round(sum_cuts, 3), round(kerf_used, 3), round(used_total, 3), round(leftover, 3), round(util, 2)])
    autosize_openpyxl_columns(ws)
    wb.save(wb_path)


def write_procurement_sheet(wb_path: str, df_summary: pd.DataFrame) -> None:
    if df_summary.empty:
        return
    agg = (
        df_summary.groupby(["Material", "Stock Length"], as_index=False)
        .agg({"Bars Used": "sum", "Estimated Waste": "sum"})
        .sort_values(["Material", "Stock Length"])
    )
    wb = load_workbook(wb_path)
    if "Procurement" in wb.sheetnames:
        del wb["Procurement"]
    ws = wb.create_sheet("Procurement")
    ws.append(list(agg.columns))
    for _, row in agg.iterrows():
        ws.append([row[c] for c in agg.columns])
    autosize_openpyxl_columns(ws)
    wb.save(wb_path)


def write_issues_sheet(wb_path: str, issues: List[dict]) -> None:
    if not issues:
        return
    wb = load_workbook(wb_path)
    if "Issues" in wb.sheetnames:
        del wb["Issues"]
    ws = wb.create_sheet("Issues")
    ws.append(["Material", "Stock Length", "Cut Length", "Problem"])
    for it in issues:
        ws.append([it["Material"], it["Stock Length"], it["Cut Length"], it["Problem"]])
    autosize_openpyxl_columns(ws)
    wb.save(wb_path)

# ---------------------------- Settings --------------------------------------

def load_settings() -> dict:
    try:
        if os.path.exists(SETTINGS_PATH):
            with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def save_settings(data: dict) -> None:
    try:
        os.makedirs(VENDOR_DIR, exist_ok=True)
        with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass

# ---------------------------- Sample Data -----------------------------------

def generate_sample_df(n: int = 50) -> pd.DataFrame:
    random.seed(404)
    materials = ["Steel Rod", "Aluminum Pipe", "Brass Tube", "PVC", "Copper Pipe", "Stainless Bar"]
    stock_lengths = [96, 120, 144, 168]
    rows = []
    for _ in range(n):
        mat = random.choice(materials)
        L = random.choice(stock_lengths)
        # 85% under stock length, 15% a little over (to exercise Issues / split mode)
        if random.random() < 0.85:
            cut = round(random.uniform(6, L - 4), 2)
        else:
            cut = round(random.uniform(L + 1, L + 40), 2)
        rows.append({"Cut Length": cut, "Max Length": L, "Material Type": mat})
    return pd.DataFrame(rows)


def save_sample_workbook(path: str, n: int = 50) -> None:
    df = generate_sample_df(n)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sample")
    wb = load_workbook(path)
    ws = wb.active
    autosize_openpyxl_columns(ws)
    wb.save(path)

# ---------------------------- GUI -------------------------------------------

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(QSize(960, 620))
        self._last_output_path = None
        self._setup_ui()
        self._load_settings()
        # default theme if not in settings
        if getattr(self, "_theme", "dark") == "dark":
            self.apply_dark_theme()
        else:
            self.apply_light_theme()

    # ---- Themes ----
    def apply_dark_theme(self) -> None:
        app = QApplication.instance()
        app.setStyle("Fusion")
        dark_palette = QPalette()
        dark_color = QColor(45, 45, 45)
        nearly_black = QColor(30, 30, 30)
        highlight = QColor(0, 122, 204)
        text_color = QColor(220, 220, 220)

        dark_palette.setColor(QPalette.Window, dark_color)
        dark_palette.setColor(QPalette.WindowText, text_color)
        dark_palette.setColor(QPalette.Base, nearly_black)
        dark_palette.setColor(QPalette.AlternateBase, dark_color)
        dark_palette.setColor(QPalette.ToolTipBase, text_color)
        dark_palette.setColor(QPalette.ToolTipText, text_color)
        dark_palette.setColor(QPalette.Text, text_color)
        dark_palette.setColor(QPalette.Button, dark_color)
        dark_palette.setColor(QPalette.ButtonText, text_color)
        dark_palette.setColor(QPalette.BrightText, QColor(255, 0, 0))
        dark_palette.setColor(QPalette.Highlight, highlight)
        dark_palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255))
        app.setPalette(dark_palette)
        self._theme = "dark"
        self._save_settings()

    def apply_light_theme(self) -> None:
        app = QApplication.instance()
        app.setStyle("Fusion")
        app.setPalette(app.style().standardPalette())  # real light palette
        self._theme = "light"
        self._save_settings()

    # ---- UI ----
    def _setup_ui(self) -> None:
        # Set window/app icon (in-app); exe icon comes from PyInstaller --icon
        logo_png = resource_path("assets/logo.png")
        if Path(logo_png).exists():
            self.setWindowIcon(QIcon(logo_png))

        central = QWidget()
        self.setCentralWidget(central)

        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        # Branding header
        header = self._make_header(logo_png)
        main_layout.addWidget(header)

        # Form box
        form_box = QFrame()
        form_box.setFrameShape(QFrame.StyledPanel)
        form_layout = QFormLayout(form_box)
        form_layout.setLabelAlignment(Qt.AlignRight)

        # File row with Browse + Create Template
        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("Select an Excel .xlsx file… or click 'Create Template'")
        browse_btn = QPushButton("Browse…")
        browse_btn.clicked.connect(self.browse_excel)
        tmpl_btn_inline = QPushButton("Create Template")
        tmpl_btn_inline.clicked.connect(self.make_template)
        path_row = QHBoxLayout()
        path_row.addWidget(self.path_edit)
        path_row.addWidget(browse_btn)
        path_row.addWidget(tmpl_btn_inline)
        form_layout.addRow("Excel File:", self._wrap(path_row))

        # Kerf
        self.kerf_edit = QLineEdit()
        self.kerf_edit.setPlaceholderText("e.g., 1/8 or 0.125")
        form_layout.addRow("Saw Kerf (inches):", self.kerf_edit)

        # Options
        self.chk_overwrite = QCheckBox("Overwrite existing 'Grouped Cuts' sheet if present")
        self.chk_copy = QCheckBox("Write to a timestamped copy of the workbook")

        # Optional reports
        self.chk_summary = QCheckBox("Add Summary sheet")
        self.chk_validation = QCheckBox("Add Validation sheet")
        self.chk_waste = QCheckBox("Add Waste Report sheet")
        self.chk_allow_split = QCheckBox("Allow oversize cut splitting (multi-piece)")
        self.chk_summary.setChecked(True)

        opts_col = QVBoxLayout()
        opts_col.addWidget(self.chk_overwrite)
        opts_col.addWidget(self.chk_copy)
        opts_col.addWidget(self._hline())
        opts_col.addWidget(QLabel("Optional reports:"))
        opts_col.addWidget(self.chk_summary)
        opts_col.addWidget(self.chk_validation)
        opts_col.addWidget(self.chk_waste)
        opts_col.addWidget(self.chk_allow_split)
        form_layout.addRow("Options:", self._wrap(opts_col))

        main_layout.addWidget(form_box)

        # Buttons row
        btn_row = QHBoxLayout()
        run_btn = QPushButton("Run Optimizer")
        run_btn.clicked.connect(self.run_optimizer)
        theme_btn = QPushButton("Toggle Theme (Light/Dark)")
        theme_btn.clicked.connect(self.toggle_theme)
        self.open_out_btn = QPushButton("Open Output File")
        self.open_out_btn.setEnabled(False)
        self.open_out_btn.clicked.connect(self.open_output_file)

        btn_row.addWidget(run_btn)
        btn_row.addWidget(self.open_out_btn)
        btn_row.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        btn_row.addWidget(theme_btn)
        main_layout.addLayout(btn_row)

        # Log / Status
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMinimumHeight(220)
        self.log.setPlaceholderText("Status messages will appear here…")
        main_layout.addWidget(self.log)

        # Menu
        menubar = self.menuBar()
        file_menu = menubar.addMenu("File")
        act_browse = QAction("Open…", self)
        act_browse.triggered.connect(self.browse_excel)
        file_menu.addAction(act_browse)

        act_create_tmpl = QAction("Create Template", self)
        act_create_tmpl.triggered.connect(self.make_template)
        file_menu.addAction(act_create_tmpl)

        act_sample = QAction("Generate Sample Data (50 rows)", self)
        act_sample.triggered.connect(self.make_sample_data)
        file_menu.addAction(act_sample)

        help_menu = menubar.addMenu("Help")
        about_action = QAction("About", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def _make_header(self, logo_png: str) -> QWidget:
        header = QWidget()
        layout = QHBoxLayout(header)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        logo_label = QLabel()
        logo_label.setFixedHeight(48)
        logo_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        if Path(logo_png).exists():
            pm = QPixmap(logo_png)
            if not pm.isNull():
                logo_label.setPixmap(pm.scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            layout.addWidget(logo_label)
        else:
            brand = QLabel("CYBEAR404")
            brand.setStyleSheet("font-weight: 800; letter-spacing: 1px;")
            brand_font = QFont()
            brand_font.setPointSize(16)
            brand.setFont(brand_font)
            layout.addWidget(brand)

        title = QLabel(APP_TITLE)
        tfont = QFont(); tfont.setPointSize(16); tfont.setBold(True)
        title.setFont(tfont)

        subtitle = QLabel("Optimize cut plans by stock length & material — with kerf.")
        subtitle.setStyleSheet("opacity: 0.85;")

        text_col = QVBoxLayout()
        text_col.addWidget(title)
        text_col.addWidget(subtitle)

        layout.addLayout(text_col)
        layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        return header

    @staticmethod
    def _wrap(layout_like) -> QWidget:
        w = QWidget()
        l = QVBoxLayout(w)
        if isinstance(layout_like, (QHBoxLayout, QVBoxLayout)):
            l.addLayout(layout_like)
        else:
            l.addWidget(layout_like)
        l.setContentsMargins(0, 0, 0, 0)
        return w

    @staticmethod
    def _hline() -> QFrame:
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        return line

    # ---- Actions ----
    def browse_excel(self) -> None:
        start_dir = os.path.dirname(self.path_edit.text().strip()) or os.getcwd()
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel workbook", start_dir, "Excel Files (*.xlsx)")
        if path:
            self.path_edit.setText(path)
            self._save_settings()

    def append_log(self, msg: str) -> None:
        self.log.append(msg)
        self.log.ensureCursorVisible()

    def run_optimizer(self) -> None:
        path = self.path_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "Missing File", "Please select an Excel .xlsx workbook, or click 'Create Template' to make one.")
            return
        if not os.path.exists(path):
            QMessageBox.critical(self, "File Not Found", f"Path does not exist:\n{path}")
            return

        try:
            kerf = parse_fraction_or_decimal(self.kerf_edit.text())
        except ValueError as e:
            QMessageBox.critical(self, "Kerf Error", str(e))
            return
        if kerf < 0:
            QMessageBox.critical(self, "Kerf Error", "Kerf cannot be negative.")
            return

        # Try open for r+ to check if file is not locked by Excel
        try:
            with open(path, "r+"):
                pass
        except Exception:
            QMessageBox.critical(self, "File Locked", "The workbook appears to be open in another program. Please close it and try again.")
            return

        self.append_log(f"Reading workbook: {path}")
        try:
            df = pd.read_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "Read Error", f"Failed to read Excel workbook.\n{e}")
            return

        required = {"Cut Length", "Max Length", "Material Type"}
        if not required.issubset(df.columns):
            missing = required - set(df.columns)
            QMessageBox.critical(self, "Missing Columns", f"The workbook must include columns: {sorted(required)}\nMissing: {sorted(missing)}")
            return

        self.append_log(f"Running optimizer (kerf={kerf})…")
        try:
            grouped, issues = bin_packing_grouped(df, kerf, allow_split_oversize=self.chk_allow_split.isChecked())
        except Exception as e:
            QMessageBox.critical(self, "Optimization Error", f"An error occurred while optimizing cuts.\n{e}")
            return

        self.append_log("Writing grouped cuts…")
        try:
            out_path = write_grouped_bins(
                path,
                grouped,
                base_sheet_name=DEFAULT_SHEET_NAME,
                overwrite_sheet=self.chk_overwrite.isChecked(),
                write_to_copy=self.chk_copy.isChecked(),
            )
        except Exception as e:
            QMessageBox.critical(self, "Write Error", f"Failed to write results.\n{e}")
            return

        # Optional reports
        try:
            if self.chk_summary.isChecked():
                self.append_log("Adding Summary sheet…")
                df_summary = build_summary(grouped, kerf)
                write_summary_sheet(out_path, df_summary)
                self.append_log("Adding Procurement sheet…")
                write_procurement_sheet(out_path, df_summary)
            if self.chk_validation.isChecked():
                self.append_log("Adding Validation sheet…")
                df_summary = df_summary if 'df_summary' in locals() else build_summary(grouped, kerf)
                write_validation_sheet(out_path, grouped, kerf, df_summary)
            if self.chk_waste.isChecked():
                self.append_log("Adding Waste Report sheet…")
                write_waste_report_sheet(out_path, grouped, kerf)
            if issues:
                self.append_log("Recording Issues (oversize cuts)…")
                write_issues_sheet(out_path, issues)
        except Exception as e:
            QMessageBox.critical(self, "Report Error", f"Failed to write optional report(s).\n{e}")
            return

        self._last_output_path = out_path
        self.open_out_btn.setEnabled(True)
        self.append_log(f"Done. Results saved to: {out_path}")
        QMessageBox.information(self, "Success", f"Optimization complete. Results saved to:\n{out_path}\n\nSheet: '{DEFAULT_SHEET_NAME}' (reports added as selected).")
        self._save_settings()

    def open_output_file(self) -> None:
        if not self._last_output_path or not os.path.exists(self._last_output_path):
            QMessageBox.warning(self, "No Output", "No output file is available yet.")
            return
        try:
            if platform.system() == "Windows":
                os.startfile(self._last_output_path)  # type: ignore[attr-defined]
            elif platform.system() == "Darwin":
                subprocess.call(["open", self._last_output_path])
            else:
                subprocess.call(["xdg-open", self._last_output_path])
        except Exception as e:
            QMessageBox.critical(self, "Open Error", f"Could not open the output file.\n{e}")

    def make_template(self) -> None:
        try:
            path = create_template_here()
            self.append_log(f"Template created: {path}")
            QMessageBox.information(self, "Template Created", f"Template saved here:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Template Error", f"Failed to create template.\n{e}")

    def make_sample_data(self) -> None:
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            name = f"autocut_sample_{ts}.xlsx"
            out_path = os.path.join(os.getcwd(), name)
            save_sample_workbook(out_path, 50)
            self.append_log(f"Sample data (50 rows) created: {out_path}")
            QMessageBox.information(self, "Sample Data", f"Sample workbook saved to:\n{out_path}")
        except Exception as e:
            QMessageBox.critical(self, "Sample Error", f"Failed to create sample workbook.\n{e}")

    def toggle_theme(self) -> None:
        if getattr(self, "_theme", "dark") == "dark":
            self.apply_light_theme()
        else:
            self.apply_dark_theme()

    def show_about(self) -> None:
        QMessageBox.information(
            self,
            "About",
            (
                f"{APP_TITLE}\n"
                f"Version: {__version__}\n\n"
                "A free proof-of-concept by Cybear404, LLC.\n"
                "Read/Write: Excel .xlsx\n"
                "Tech: PySide6 (Qt), pandas, openpyxl\n\n"
                "Input columns required in the first sheet:\n"
                " - Cut Length (numeric)\n"
                " - Max Length (numeric)\n"
                " - Material Type (text)\n\n"
                "License: PolyForm Noncommercial 1.0.0 (noncommercial use only)\n"
                "Commercial licenses available — see COMMERCIAL-LICENSE.md\n"
                "Branding: Cybear404 name/logo are proprietary — see BRANDING.md\n"
            ),
        )

    # ---- Settings bind/load/save ----
    def _collect_settings(self) -> dict:
        return {
            "last_path": self.path_edit.text().strip(),
            "kerf_text": self.kerf_edit.text().strip(),
            "overwrite": self.chk_overwrite.isChecked(),
            "copy": self.chk_copy.isChecked(),
            "summary": self.chk_summary.isChecked(),
            "validation": self.chk_validation.isChecked(),
            "waste": self.chk_waste.isChecked(),
            "allow_split": self.chk_allow_split.isChecked(),
            "theme": getattr(self, "_theme", "dark"),
            "last_output": self._last_output_path or "",
        }

    def _apply_settings(self, s: dict) -> None:
        self.path_edit.setText(s.get("last_path", ""))
        self.kerf_edit.setText(s.get("kerf_text", ""))
        self.chk_overwrite.setChecked(bool(s.get("overwrite", False)))
        self.chk_copy.setChecked(bool(s.get("copy", False)))
        self.chk_summary.setChecked(bool(s.get("summary", True)))
        self.chk_validation.setChecked(bool(s.get("validation", False)))
        self.chk_waste.setChecked(bool(s.get("waste", False)))
        self.chk_allow_split.setChecked(bool(s.get("allow_split", False)))
        self._theme = s.get("theme", "dark")
        self._last_output_path = s.get("last_output") or None
        self.open_out_btn.setEnabled(bool(self._last_output_path and os.path.exists(self._last_output_path)))

    def _load_settings(self) -> None:
        s = load_settings()
        self._apply_settings(s)

    def _save_settings(self) -> None:
        s = self._collect_settings()
        save_settings(s)

    def closeEvent(self, event):
        self._save_settings()
        super().closeEvent(event)

# ---------------------------- Template helper -------------------------------

def create_template_here() -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"{TEMPLATE_BASENAME}_{ts}.xlsx"
    path = os.path.join(os.getcwd(), name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Cut Length", "Max Length", "Material Type"])
    ws.append([100, 120, "Steel Rod"])
    ws.append([50, 120, "Steel Rod"])
    ws.append([75, 144, "Aluminum Pipe"])
    autosize_openpyxl_columns(ws)

    wb.save(path)
    wb.close()
    return path

# ---------------------------- Entry point -----------------------------------

def main() -> None:
    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    # Also set app icon here (affects taskbar/dock in some environments)
    logo_png = resource_path("assets/logo.png")
    if Path(logo_png).exists():
        app.setWindowIcon(QIcon(logo_png))

    win = MainWindow()
    win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()